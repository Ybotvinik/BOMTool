"""Parse uploaded BOM tables (Excel .xlsx/.xls or CSV).

Customer BOM files frequently contain metadata rows (Board Name, Doc. Number,
Revised Date, blank rows) before the real header row, so we never assume row 0
is the header — we scan the first rows and score candidate header rows.
"""

from __future__ import annotations

import csv
import io
import re

from openpyxl import load_workbook

# Canonical BOM line fields the importer can map columns onto. Order matters for
# UI display only; mapping itself is conflict-aware (see ``suggest_mapping``).
BOM_FIELDS: list[str] = [
    "mpn",
    "manufacturer",
    "description",
    "quantity",
    "reference_designators",
    "footprint",
    "value",
    "supplier_part_number",
    "unit",
    "customer_price",
    "internal_cost",
    "is_critical",
    "dnp",
]

# Keywords that indicate a row is a real BOM header row.
HEADER_KEYWORDS: list[str] = [
    "quantity",
    "qty",
    "designator",
    "refdes",
    "reference",
    "part number",
    "part no",
    "manufacturer part number",
    "mpn",
    "manufacturer",
    "description",
    "value",
    "footprint",
    "package",
    "supplier part number",
    "datasheet",
    "assembly",
    "dnp",
    "comment",
    "p/n",
]

# Per-field hints (longer / more specific hints win during assignment).
FIELD_HINTS: dict[str, list[str]] = {
    "mpn": [
        "manufacturer part number",
        "mfr part number",
        "mfg part number",
        "mpn",
        "part number",
        "part no",
        "p/n",
        "pn",
        'מק"ט',
        "מקט",
    ],
    "manufacturer": ["manufacturer", "mfr", "mfg", "maker", "brand", "יצרן"],
    "description": ["part description", "description", "comment", "desc", "details", "תיאור"],
    "quantity": ["quantity", "qty", "amount", "כמות"],
    "reference_designators": [
        "reference designator",
        "designator",
        "refdes",
        "references",
        "reference",
        "ref des",
    ],
    "footprint": ["pcb footprint", "footprint", "package", "pkg"],
    "value": ["value", "ערך"],
    "supplier_part_number": [
        "supplier part number",
        "supplier p/n",
        "supplier pn",
        "digi-key pn",
        "digikey pn",
        "digi-key part number",
        "mouser pn",
        "mouser part number",
    ],
    "unit": ["unit", "uom", "יחידה"],
    "customer_price": ["customer price", "sell price", "price", "מחיר ללקוח", "מחיר"],
    "internal_cost": ["internal cost", "unit cost", "cost", "עלות"],
    "is_critical": ["critical", "קריטי"],
    "dnp": ["dnp", "dni", "do not populate", "do not place", "assembly", "fitted", "populate", "mount"],
}

SCAN_ROWS = 30


def normalize_key(name: str) -> str:
    """Lowercase, strip, drop line breaks, collapse whitespace — for matching."""
    s = (name or "").replace("\n", " ").replace("\r", " ").strip().lower()
    return re.sub(r"\s+", " ", s)


def clean_display(name: str) -> str:
    """Trim + collapse whitespace but keep the original casing for display."""
    s = (name or "").replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", s)


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def list_sheet_names(content: bytes, filename: str) -> tuple[list[str], str]:
    """Return (sheet_names, active_sheet_name) without loading row data."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return ["CSV"], "CSV"

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    active = wb.active.title if wb.active else (sheet_names[0] if sheet_names else "")
    wb.close()
    return sheet_names, active


def list_sheets_and_rows(
    content: bytes, filename: str, sheet_name: str | None = None
) -> tuple[list[str], str, list[list[str]]]:
    """Return (sheet_names, active_sheet_name, rows) for the chosen sheet."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        rows = [[_stringify(c) for c in r] for r in csv.reader(io.StringIO(text))]
        return ["CSV"], "CSV", rows

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    ws = wb[sheet_name] if sheet_name in sheet_names else wb.active
    rows = [[_stringify(c) for c in raw] for raw in ws.iter_rows(values_only=True)]
    wb.close()
    return sheet_names, ws.title, rows


def _keyword_hits(row: list[str]) -> int:
    norm_cells = [normalize_key(c) for c in row if c.strip()]
    return sum(1 for c in norm_cells if any(kw in c for kw in HEADER_KEYWORDS))


def detect_header_row(rows: list[list[str]]) -> tuple[int | None, list[dict]]:
    """Scan the first rows and return (detected_index, candidate_rows).

    A candidate header row has >= 3 non-empty cells. The detected row is the
    candidate with the most BOM-keyword hits (requires >= 2 to be considered a
    real header), tie-broken by more non-empty cells then earliest row.
    """
    candidates: list[dict] = []
    for i, row in enumerate(rows[:SCAN_ROWS]):
        non_empty = [c for c in row if c.strip()]
        if len(non_empty) < 3:
            continue
        candidates.append(
            {
                "row_index": i,
                "values": [clean_display(c) for c in row],
                "non_empty_count": len(non_empty),
                "keyword_hits": _keyword_hits(row),
            }
        )

    detected: int | None = None
    best: dict | None = None
    for c in candidates:
        if c["keyword_hits"] < 2:
            continue
        if (
            best is None
            or c["keyword_hits"] > best["keyword_hits"]
            or (
                c["keyword_hits"] == best["keyword_hits"]
                and c["non_empty_count"] > best["non_empty_count"]
            )
        ):
            best = c
    if best is not None:
        detected = best["row_index"]
    return detected, candidates


def _best_hint_score(col_norm: str, hints: list[str]) -> int:
    best = 0
    for hint in hints:
        hn = normalize_key(hint)
        if hn and hn in col_norm:
            score = len(hn) * (3 if col_norm == hn else 1)
            best = max(best, score)
    return best


def suggest_mapping_generic(
    columns: list[str], field_hints: dict[str, list[str]]
) -> dict[str, str | None]:
    """Conflict-aware greedy mapping of columns to fields by hint score.

    Longer/more specific hints win; each column maps to at most one field.
    """
    cols = [(c, normalize_key(c)) for c in columns if c.strip()]
    pairs: list[tuple[int, str, str]] = []
    for field, hints in field_hints.items():
        for disp, cn in cols:
            score = _best_hint_score(cn, hints)
            if score > 0:
                pairs.append((score, field, disp))
    pairs.sort(key=lambda x: -x[0])

    assigned_field: dict[str, str] = {}
    used_cols: set[str] = set()
    for _, field, disp in pairs:
        if field in assigned_field or disp in used_cols:
            continue
        assigned_field[field] = disp
        used_cols.add(disp)
    return {f: assigned_field.get(f) for f in field_hints}


def suggest_mapping(columns: list[str]) -> dict[str, str | None]:
    """Conflict-aware auto mapping of columns to BOM fields."""
    mapping = suggest_mapping_generic(columns, FIELD_HINTS)
    mapping = {f: mapping.get(f) for f in BOM_FIELDS}
    if mapping.get("mpn") is None and mapping.get("supplier_part_number"):
        mapping["mpn"] = mapping["supplier_part_number"]
    return mapping


# --- BOM revision / metadata detection ---------------------------------------

# Field -> ordered label variants (most specific first). Order across fields
# matters: "revised date" is matched before "revision"/"rev".
_METADATA_LABELS: list[tuple[str, list[str]]] = [
    ("board_name", ["board name", "board", "pcb name", "assembly name"]),
    ("doc_number", ["document number", "doc. number", "doc number", "doc no", "document no"]),
    ("bom_number", ["bom number", "bom no", "bom #"]),
    ("revised_date", ["revised date", "release date", "rev date", "date"]),
    ("revision", ["revision", "rev."]),
    ("revision", ["rev"]),
]

_REVISION_RE = re.compile(r"(?i)(?:^|[-_\s:])R\s*-?\s*(\d+[a-z]?)\b")
_DATE_FORMATS = [
    "%d.%m.%Y",
    "%d.%m.%y",
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
    "%d-%m-%Y",
]


def extract_revision(text: str) -> str | None:
    """Pull a revision token like ``R09`` out of a string (e.g. a doc number)."""
    if not text:
        return None
    m = _REVISION_RE.search(text)
    if m:
        return f"R{m.group(1).upper()}"
    return None


def _normalize_date(text: str) -> str | None:
    from datetime import datetime

    raw = (text or "").strip()
    if not raw:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _match_label(cell_norm: str) -> str | None:
    for field, labels in _METADATA_LABELS:
        for lab in labels:
            if cell_norm == lab or cell_norm.startswith(lab + " "):
                return field
    return None


def extract_metadata(rows: list[list[str]], header_index: int | None) -> dict[str, str | None]:
    """Extract board/doc/revision/date metadata from rows above the header.

    Handles 'label: value' in one cell, and 'label' + adjacent-cell value.
    """
    out: dict[str, str | None] = {
        "board_name": None,
        "doc_number": None,
        "revised_date": None,
        "revision": None,
        "bom_number": None,
    }
    upper = header_index if header_index is not None else min(len(rows), SCAN_ROWS)
    for row in rows[:upper]:
        for i, cell in enumerate(row):
            text = clean_display(cell)
            if not text:
                continue
            # Case A: "label: value" in a single cell.
            if ":" in text:
                label_part, _, value_part = text.partition(":")
                field = _match_label(normalize_key(label_part))
                if field and out.get(field) is None and value_part.strip():
                    out[field] = value_part.strip()
                    continue
            # Case B: cell is a label; value is the next non-empty cell.
            field = _match_label(normalize_key(text))
            if field and out.get(field) is None:
                for nxt in row[i + 1 :]:
                    val = clean_display(nxt)
                    if val:
                        out[field] = val
                        break

    # Derive a clean revision code from revision text or the doc number.
    rev_code: str | None = None
    if out["revision"]:
        rev_code = extract_revision(out["revision"]) or _format_rev(out["revision"])
    if not rev_code and out["doc_number"]:
        rev_code = extract_revision(out["doc_number"])
    out["revision_code"] = rev_code
    return out


def _format_rev(raw: str) -> str | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    if raw.isdigit():
        return f"R{raw.zfill(2)}"
    return raw


def suggest_version_name(
    metadata: dict[str, str | None], existing_names: list[str]
) -> tuple[str, str | None]:
    """Return (suggested_version_name, revision_code) following the spec rules."""
    existing = set(existing_names)
    revision_code = metadata.get("revision_code")

    base: str | None = revision_code
    if not base:
        date_iso = _normalize_date(metadata.get("revised_date") or "")
        if date_iso:
            base = f"v{date_iso}"
    if not base:
        # Next sequential vN based on existing v-numbers.
        max_n = 0
        for name in existing_names:
            m = re.fullmatch(r"v(\d+)", name.strip())
            if m:
                max_n = max(max_n, int(m.group(1)))
        base = f"v{max_n + 1}"

    name = base
    if name in existing:
        n = 2
        while f"{base}-{n}" in existing:
            n += 1
        name = f"{base}-{n}"
    return name, revision_code
