"""Excel export builders for customer-safe and internal reports."""

from __future__ import annotations

import io
import re
from copy import copy
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import BomLine, BomVersion, Customer, OfficialPriceLine, PricingLine, PricingSnapshot, Project
from app.services.suppliers.official_pricing import (
    get_latest_exportable_snapshot,
    snapshot_lines_by_bom_line,
)

# Official/market prices in this export are reference prices only — not final customer
# quote prices. Customer Quote Price is a future commercial pricing layer.
# Internal/China costs must never be exposed in customer exports.
# Official Unit Price is populated only when notes contain an explicit official source.

CUSTOMER_TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "templates" / "customer_bom_cost_review_template.xlsx"
)
INTERNAL_TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent
    / "templates"
    / "internal_reports"
    / "Internal Pricing Snapshot Excel.xlsx"
)
INTERNAL_TEMPLATE_SHEET = "Pricing Lines"
INTERNAL_WARNING_ROW = 1
INTERNAL_TITLE_ROW = 2
INTERNAL_META_START_ROW = 4
INTERNAL_TABLE_HEADER_ROW = 9
INTERNAL_DATA_START_ROW = 10
INTERNAL_TEMPLATE_META_ROW = 3
INTERNAL_TEMPLATE_HEADER_ROW = 8
INTERNAL_TEMPLATE_DATA_ROW = 9
INTERNAL_TEMPLATE_COLS = 20
INTERNAL_BANNER_MERGE_COLS = 20
CUSTOMER_SHEET_NAME = "Customer BOM Cost Review"
TEMPLATE_TABLE_HEADER_ROW = 15
TEMPLATE_DATA_START_ROW = 16
TEMPLATE_DATA_DEFAULT_END_ROW = 90
TEMPLATE_TABLE_COLS = 12

_COL_LINE = 1
_COL_MPN = 2
_COL_MANUFACTURER = 3
_COL_DESCRIPTION = 4
_COL_QTY = 5
_COL_REQ_QTY = 6
_COL_REFDES = 7
_COL_FOOTPRINT = 8
_COL_OFFICIAL_SOURCE = 9
_COL_OFFICIAL_UNIT = 10
_COL_OFFICIAL_EXTENDED = 11
_COL_NOTES = 12

# Backward-compatible aliases used by column index logic.
_COL_PRICE_SOURCE = _COL_OFFICIAL_SOURCE
_COL_UNIT_PRICE = _COL_OFFICIAL_UNIT
_COL_EXTENDED = _COL_OFFICIAL_EXTENDED

USD_CURRENCY_FMT = '"$"#,##0.00'
USD_UNIT_FMT = '"$"#,##0.0000'
QTY_INT_FMT = "#,##0"

_METADATA_BUILD_QTY = "G9"

_DNP_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
_DNP_FONT_COLOR = "808080"
_DNP_NOTE = "DNP / Not populated"

# --- Internal report styling (customer export must not use these) ---
_INTERNAL_WARNING_EN = "GLINTECH INTERNAL ONLY"
_INTERNAL_WARNING_HE = "פנימי בלבד — לא להעברה ללקוח"

INTERNAL_REPORT_STYLES: dict[str, dict[str, str]] = {
    "bom_quality": {
        "title_en": "GLINTECH INTERNAL BOM QUALITY REPORT",
        "title_he": "דוח איכות BOM פנימי",
        "color": "1F4E79",
    },
    "pricing_snapshot": {
        "title_en": "GLINTECH INTERNAL PRICING SNAPSHOT",
        "title_he": "תמונת מצב מחירים פנימית",
        "color": "7030A0",
    },
    "pricing_comparison": {
        "title_en": "GLINTECH INTERNAL PRICING COMPARISON",
        "title_he": "השוואת מחירים פנימית",
        "color": "548235",
    },
    "supplier_purchase": {
        "title_en": "GLINTECH SUPPLIER PURCHASE REPORT",
        "title_he": "דוח רכש ספקים פנימי",
        "color": "C65911",
    },
    "summary": {
        "title_en": "GLINTECH INTERNAL SUMMARY",
        "title_he": "סיכום פנימי",
        "color": "1F3864",
    },
}

_INTERNAL_REPORT_TYPE_ALIASES: dict[str, str] = {
    "quality": "bom_quality",
    "pricing": "pricing_snapshot",
    "comparison": "pricing_comparison",
    "purchase": "supplier_purchase",
}

_THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
_ALIGN_RTL_TEXT = Alignment(horizontal="right", vertical="center", wrap_text=True)
_ALIGN_RTL_NUMBER = Alignment(horizontal="left", vertical="center")
_NEUTRAL_DATA_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
_NEUTRAL_DATA_FONT = Font(color="000000", size=11)

_FORBIDDEN_CUSTOMER_HEADER_PHRASES = (
    "china",
    "link",
    "east supplier",
    "internal",
    "margin",
    "savings",
    "confidence",
    "buyer",
    "match confidence",
    "internal cost",
    "china cost",
    "unit cost",
    "extended cost",
    "gross margin",
    "customer unit price",
    "customer extended price",
    "customer price source",
    "customer price list",
    "official price list",
    "generic price list",
)

# (needle, canonical label) — longest needles first for matching in BOM line notes.
_OFFICIAL_SOURCE_PATTERNS: tuple[tuple[str, str], ...] = (
    ("future electronics", "Future"),
    ("manual official", "Manual Official"),
    ("official rep", "Official Rep"),
    ("digi-key", "Digi-Key"),
    ("digikey", "Digi-Key"),
    ("texas instruments", "TI"),
    ("mouser", "Mouser"),
    ("ti", "TI"),
    ("arrow", "Arrow"),
    ("avnet", "Avnet"),
    ("future", "Future"),
    ("tti", "TTI"),
)

_FORBIDDEN_OFFICIAL_SOURCE_VALUES = frozenset(
    {
        "official price list",
        "customer price list",
        "generic price list",
        "price list",
        "official/market price list",
        "link",
        "china",
        "east",
    }
)


def _safe_part(value: str | None) -> str:
    if not value:
        return "unknown"
    return re.sub(r"[^\w\-]+", "_", str(value).strip()).strip("_") or "unknown"


def _num(value: Decimal | float | int | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _autosize(ws, *, min_col: int = 1, max_col: int | None = None) -> None:
    last_col = max_col or ws.max_column
    for col_idx in range(min_col, last_col + 1):
        column_cells = [ws.cell(row=r, column=col_idx) for r in range(1, ws.max_row + 1)]
        max_len = 0
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        if max_len:
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def _load_internal_template_workbook() -> Workbook:
    if not INTERNAL_TEMPLATE_PATH.is_file():
        raise FileNotFoundError(
            f"Internal report template not found: {INTERNAL_TEMPLATE_PATH}"
        )
    return load_workbook(INTERNAL_TEMPLATE_PATH)


def _internal_report_style(report_type: str) -> dict[str, str]:
    key = _INTERNAL_REPORT_TYPE_ALIASES.get(report_type, report_type)
    if key not in INTERNAL_REPORT_STYLES:
        raise ValueError(f"Unknown internal report type: {report_type}")
    return INTERNAL_REPORT_STYLES[key]


def apply_rtl(ws) -> None:
    ws.sheet_view.rightToLeft = True


def apply_internal_warning_banner(ws, span_cols: int | None = None) -> None:
    apply_rtl(ws)
    span = max(span_cols or INTERNAL_BANNER_MERGE_COLS, 4)
    merge_ref = (
        f"A{INTERNAL_WARNING_ROW}:{get_column_letter(span)}{INTERNAL_WARNING_ROW}"
    )
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row == INTERNAL_WARNING_ROW:
            ws.unmerge_cells(str(merged))
    ws.merge_cells(merge_ref)
    cell = ws.cell(
        row=INTERNAL_WARNING_ROW,
        column=1,
        value=f"{_INTERNAL_WARNING_EN}  |  {_INTERNAL_WARNING_HE}",
    )
    tpl_wb = _load_internal_template_workbook()
    tpl_cell = tpl_wb.active.cell(row=INTERNAL_TEMPLATE_META_ROW - 2, column=1)
    if tpl_cell.fill and tpl_cell.fill.fill_type:
        cell.fill = copy(tpl_cell.fill)
    else:
        cell.fill = PatternFill(fill_type="solid", fgColor="FEF3C7")
    cell.font = Font(bold=True, color="991B1B", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[INTERNAL_WARNING_ROW].height = 28


def apply_internal_report_title(ws, report_type: str, span_cols: int | None = None) -> None:
    style = _internal_report_style(report_type)
    span = max(span_cols or INTERNAL_BANNER_MERGE_COLS, 4)
    merge_ref = f"A{INTERNAL_TITLE_ROW}:{get_column_letter(span)}{INTERNAL_TITLE_ROW}"
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row == INTERNAL_TITLE_ROW:
            ws.unmerge_cells(str(merged))
    ws.merge_cells(merge_ref)
    cell = ws.cell(
        row=INTERNAL_TITLE_ROW,
        column=1,
        value=f"{style['title_en']}\n{style['title_he']}",
    )
    cell.fill = PatternFill(fill_type="solid", fgColor=style["color"])
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[INTERNAL_TITLE_ROW].height = 44


def apply_internal_report_layout(
    ws,
    report_type: str,
    *,
    span_cols: int | None = None,
) -> None:
    span = max(span_cols or INTERNAL_BANNER_MERGE_COLS, 4)
    apply_rtl(ws)
    apply_internal_warning_banner(ws, span_cols=span)
    apply_internal_report_title(ws, report_type, span_cols=span)


def _apply_template_column_widths(ws, tpl_ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        letter = get_column_letter(col)
        width = tpl_ws.column_dimensions[letter].width
        if width:
            ws.column_dimensions[letter].width = width


def style_internal_headers(
    ws, row: int, num_cols: int, report_type: str, tpl_ws
) -> None:
    color = _internal_report_style(report_type)["color"]
    fill = PatternFill(fill_type="solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        tpl_hdr = tpl_ws.cell(
            row=INTERNAL_TEMPLATE_HEADER_ROW, column=min(col, INTERNAL_TEMPLATE_COLS)
        )
        cell.fill = fill
        cell.font = font
        cell.border = copy(tpl_hdr.border) if tpl_hdr.border else _THIN_BORDER
        if tpl_hdr.alignment:
            cell.alignment = copy(tpl_hdr.alignment)
        else:
            cell.alignment = _ALIGN_RTL_TEXT


def _clear_internal_values(ws, num_cols: int) -> None:
    for row in range(INTERNAL_META_START_ROW, ws.max_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = _NEUTRAL_DATA_FILL
            cell.font = _NEUTRAL_DATA_FONT


def _apply_neutral_data_cell_style(cell, val, tpl_cell) -> None:
    """Data rows: neutral fill/font only — never report header colors."""
    cell.fill = _NEUTRAL_DATA_FILL
    cell.font = _NEUTRAL_DATA_FONT
    cell.border = copy(tpl_cell.border) if tpl_cell.border else _THIN_BORDER
    if tpl_cell.alignment:
        cell.alignment = copy(tpl_cell.alignment)
    elif isinstance(val, (int, float)):
        cell.alignment = _ALIGN_RTL_NUMBER
    else:
        cell.alignment = _ALIGN_RTL_TEXT


def _clear_stale_internal_data_rows(
    ws, num_cols: int, last_data_row: int
) -> None:
    """Remove leftover template rows below the written data block."""
    for row in range(last_data_row + 1, ws.max_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = _NEUTRAL_DATA_FILL
            cell.font = _NEUTRAL_DATA_FONT


def _write_internal_meta_template(
    ws,
    rows: list[tuple[str, object]],
    tpl_ws,
    *,
    kpi_rows: list[tuple[str, object]] | None = None,
) -> None:
    bold = Font(bold=True)
    for idx, (label, val) in enumerate(rows[:4]):
        row = INTERNAL_META_START_ROW + idx
        tpl_row = INTERNAL_TEMPLATE_META_ROW + idx
        label_cell = ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=2, value=val)
        tpl_label = tpl_ws.cell(row=tpl_row, column=1)
        tpl_value = tpl_ws.cell(row=tpl_row, column=2)
        label_cell.font = copy(tpl_label.font) if tpl_label.font else bold
        value_cell.font = copy(tpl_value.font) if tpl_value.font else _NEUTRAL_DATA_FONT
        label_cell.fill = _NEUTRAL_DATA_FILL
        value_cell.fill = _NEUTRAL_DATA_FILL
        label_cell.alignment = (
            copy(tpl_label.alignment) if tpl_label.alignment else _ALIGN_RTL_TEXT
        )
        value_cell.alignment = (
            copy(tpl_value.alignment) if tpl_value.alignment else Alignment(horizontal="left")
        )
    if kpi_rows:
        for idx, (label, val) in enumerate(kpi_rows[:4]):
            row = INTERNAL_META_START_ROW + idx
            tpl_row = INTERNAL_TEMPLATE_META_ROW + idx
            label_cell = ws.cell(row=row, column=4, value=label)
            value_cell = ws.cell(row=row, column=5, value=val)
            tpl_label = tpl_ws.cell(row=tpl_row, column=4)
            tpl_value = tpl_ws.cell(row=tpl_row, column=5)
            label_cell.font = copy(tpl_label.font) if tpl_label.font else bold
            value_cell.font = copy(tpl_value.font) if tpl_value.font else _NEUTRAL_DATA_FONT
            label_cell.fill = _NEUTRAL_DATA_FILL
            value_cell.fill = _NEUTRAL_DATA_FILL
            label_cell.alignment = (
                copy(tpl_label.alignment) if tpl_label.alignment else Alignment(horizontal="left")
            )
            value_cell.alignment = (
                copy(tpl_value.alignment) if tpl_value.alignment else Alignment(horizontal="center")
            )
    extra = rows[4:]
    if extra:
        row = INTERNAL_META_START_ROW + 4
        for label, val in extra:
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = bold
            label_cell.fill = _NEUTRAL_DATA_FILL
            label_cell.alignment = _ALIGN_RTL_TEXT
            value_cell = ws.cell(row=row, column=2, value=val)
            value_cell.fill = _NEUTRAL_DATA_FILL
            value_cell.font = _NEUTRAL_DATA_FONT
            value_cell.alignment = _ALIGN_RTL_TEXT
            row += 1


def _write_internal_table_data(
    ws, data_rows: list[list], num_cols: int, tpl_ws
) -> None:
    for row_offset, values in enumerate(data_rows):
        row = INTERNAL_DATA_START_ROW + row_offset
        for col, val in enumerate(values[:num_cols], start=1):
            cell = ws.cell(row=row, column=col, value=val)
            tpl_cell = tpl_ws.cell(
                row=INTERNAL_TEMPLATE_DATA_ROW, column=min(col, INTERNAL_TEMPLATE_COLS)
            )
            _apply_neutral_data_cell_style(cell, val, tpl_cell)
    if data_rows:
        last_row = INTERNAL_DATA_START_ROW + len(data_rows) - 1
    else:
        last_row = INTERNAL_DATA_START_ROW - 1
    _clear_stale_internal_data_rows(ws, num_cols, last_row)


def _render_internal_data_sheet(
    ws,
    *,
    report_type: str,
    meta_rows: list[tuple[str, object]],
    headers: list[str],
    data_rows: list[list],
    kpi_rows: list[tuple[str, object]] | None = None,
    tpl_ws=None,
) -> None:
    tpl_wb = None
    if tpl_ws is None:
        tpl_wb = _load_internal_template_workbook()
        tpl_ws = tpl_wb.active
    num_cols = max(len(headers), len(data_rows[0]) if data_rows else len(headers))
    num_cols = max(num_cols, 4)
    span = max(num_cols, INTERNAL_BANNER_MERGE_COLS)
    apply_internal_report_layout(ws, report_type, span_cols=span)
    _clear_internal_values(ws, max(num_cols, INTERNAL_TEMPLATE_COLS))
    _write_internal_meta_template(ws, meta_rows, tpl_ws, kpi_rows=kpi_rows)
    for col, header in enumerate(headers, start=1):
        ws.cell(row=INTERNAL_TABLE_HEADER_ROW, column=col, value=header)
    style_internal_headers(ws, INTERNAL_TABLE_HEADER_ROW, len(headers), report_type, tpl_ws)
    _write_internal_table_data(ws, data_rows, len(headers), tpl_ws)
    _apply_template_column_widths(ws, tpl_ws, max(len(headers), INTERNAL_TEMPLATE_COLS))
    ws.freeze_panes = ws.cell(row=INTERNAL_DATA_START_ROW, column=1)
    apply_rtl(ws)


def _render_internal_summary_sheet(
    ws,
    *,
    report_type: str,
    meta_rows: list[tuple[str, object]],
    section_title: str | None = None,
    section_rows: list[tuple[str, object]] | None = None,
    tpl_ws=None,
) -> None:
    tpl_wb = None
    if tpl_ws is None:
        tpl_wb = _load_internal_template_workbook()
        tpl_ws = tpl_wb.active
    apply_internal_report_layout(ws, report_type, span_cols=INTERNAL_BANNER_MERGE_COLS)
    _clear_internal_values(ws, INTERNAL_TEMPLATE_COLS)
    row = INTERNAL_META_START_ROW
    bold = Font(bold=True)
    for label, val in meta_rows:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = bold
        label_cell.fill = _NEUTRAL_DATA_FILL
        label_cell.alignment = _ALIGN_RTL_TEXT
        value_cell = ws.cell(row=row, column=2, value=val)
        value_cell.fill = _NEUTRAL_DATA_FILL
        value_cell.font = _NEUTRAL_DATA_FONT
        value_cell.alignment = _ALIGN_RTL_TEXT
        row += 1
    if section_title:
        row += 1
        ws.cell(row=row, column=1, value=section_title)
        style_internal_headers(ws, row, 2, "summary", tpl_ws)
        row += 1
    if section_rows:
        for label, val in section_rows:
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = bold
            label_cell.fill = _NEUTRAL_DATA_FILL
            label_cell.alignment = _ALIGN_RTL_TEXT
            value_cell = ws.cell(row=row, column=2, value=val)
            value_cell.fill = _NEUTRAL_DATA_FILL
            value_cell.font = _NEUTRAL_DATA_FONT
            value_cell.alignment = _ALIGN_RTL_NUMBER
            row += 1
    _apply_template_column_widths(ws, tpl_ws, INTERNAL_TEMPLATE_COLS)
    apply_rtl(ws)


def _create_internal_workbook(sheet_title: str) -> tuple[Workbook, object, object]:
    tpl_wb = _load_internal_template_workbook()
    tpl_ws = tpl_wb.active
    ws = tpl_ws
    ws.title = sheet_title
    return tpl_wb, ws, tpl_ws


def _add_internal_sheet(wb: Workbook, sheet_title: str, tpl_ws) -> object:
    ws = wb.copy_worksheet(wb.worksheets[0])
    ws.title = sheet_title
    return ws


def _finalize_internal_workbook(wb: Workbook) -> None:
    if wb.views:
        wb.views[0].rightToLeft = True
    for ws in wb.worksheets:
        apply_rtl(ws)


def _write_header_block(ws, rows: list[tuple[str, str]]) -> None:
    bold = Font(bold=True)
    for label, val in rows:
        ws.append([label, val])
        ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([])


def _validate_customer_export_headers(headers: list[str]) -> None:
    """Reject customer exports that accidentally include internal-only columns."""
    for header in headers:
        if not header:
            continue
        lower = str(header).lower().strip()
        if any(phrase in lower for phrase in _FORBIDDEN_CUSTOMER_HEADER_PHRASES):
            raise ValueError(
                f"Unsafe customer export header '{header}': contains forbidden internal term"
            )


def _copy_cell_style(src, dst) -> None:
    if not src.has_style:
        return
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def _snapshot_row_styles(ws, row: int, max_col: int = TEMPLATE_TABLE_COLS) -> list:
    return [ws.cell(row=row, column=c) for c in range(1, max_col + 1)]


def _apply_row_styles(ws, row: int, style_sources: list) -> None:
    for col_idx, src in enumerate(style_sources, start=1):
        _copy_cell_style(src, ws.cell(row=row, column=col_idx))


def _apply_dnp_row_styles(ws, row: int, style_sources: list) -> None:
    """Gray out entire DNP data row while preserving borders/alignment."""
    for col_idx, src in enumerate(style_sources, start=1):
        cell = ws.cell(row=row, column=col_idx)
        _copy_cell_style(src, cell)
        cell.fill = _DNP_FILL
        if cell.font:
            cell.font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=cell.font.bold,
                italic=cell.font.italic,
                underline=cell.font.underline,
                strikethrough=cell.font.strikethrough,
                color=_DNP_FONT_COLOR,
            )
        else:
            cell.font = Font(color=_DNP_FONT_COLOR)


def _is_dnp_line(line: BomLine) -> bool:
    if line.dnp:
        return True
    if line.required_qty is not None and float(line.required_qty) == 0:
        return True
    return False


def _load_customer_template_sheet():
    if not CUSTOMER_TEMPLATE_PATH.is_file():
        raise FileNotFoundError(
            f"Customer BOM template not found: {CUSTOMER_TEMPLATE_PATH}"
        )
    wb = load_workbook(CUSTOMER_TEMPLATE_PATH)
    if CUSTOMER_SHEET_NAME in wb.sheetnames:
        ws = wb[CUSTOMER_SHEET_NAME]
    else:
        ws = wb.active
    ws.sheet_view.rightToLeft = False
    return wb, ws


def _fill_customer_template_metadata(
    ws,
    *,
    customer_name: str,
    project_name: str,
    project_code: str,
    version_label: str,
    revision: str,
    doc_number: str,
    board_name: str,
    build_qty: int,
    export_date: str,
) -> None:
    """Fill project header + summary qty cells defined by the template layout."""
    ws["B4"] = customer_name
    ws["D4"] = project_name
    ws["G4"] = project_code
    ws["B5"] = version_label
    ws["D5"] = revision
    ws["G5"] = doc_number
    ws["B6"] = board_name
    ws["D6"] = build_qty if build_qty > 0 else ""
    ws["G6"] = export_date
    ws[_METADATA_BUILD_QTY] = build_qty if build_qty > 0 else ""
    if build_qty > 0:
        ws[_METADATA_BUILD_QTY].number_format = QTY_INT_FMT


def _cell_has_formula(cell) -> bool:
    value = cell.value
    return isinstance(value, str) and value.startswith("=")


def _snapshot_template_extended_formula(ws) -> str | None:
    """Capture the row-level extended price formula from the template."""
    value = ws.cell(
        row=TEMPLATE_DATA_START_ROW, column=_COL_OFFICIAL_EXTENDED
    ).value
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _apply_template_extended_formula(ws, row: int, template_formula: str) -> None:
    """Copy the template extended-price formula to another data row."""
    ext_col = get_column_letter(_COL_OFFICIAL_EXTENDED)
    origin = f"{ext_col}{TEMPLATE_DATA_START_ROW}"
    dest = f"{ext_col}{row}"
    ws.cell(row=row, column=_COL_OFFICIAL_EXTENDED).value = Translator(
        template_formula, origin=origin
    ).translate_formula(dest)


def _clear_bom_row_values(ws, start: int, end: int) -> None:
    """Clear prior BOM sample values; preserve existing extended-price formulas."""
    for row in range(start, end + 1):
        for col in range(1, TEMPLATE_TABLE_COLS + 1):
            if col == _COL_OFFICIAL_EXTENDED and _cell_has_formula(
                ws.cell(row=row, column=col)
            ):
                continue
            ws.cell(row=row, column=col).value = None


def customer_bom_review_filename(project_code: str, version_name: str) -> str:
    return f"Customer_BOM_Review_{_safe_part(project_code)}_{_safe_part(version_name)}.xlsx"


def internal_bom_quality_filename(project_code: str, version_name: str) -> str:
    return f"Internal_BOM_Quality_{_safe_part(project_code)}_{_safe_part(version_name)}.xlsx"


def internal_pricing_snapshot_filename(project_code: str, snapshot_name: str) -> str:
    return f"Internal_Pricing_{_safe_part(project_code)}_{_safe_part(snapshot_name)}.xlsx"


def _is_forbidden_official_source(value: str | None) -> bool:
    if not value:
        return False
    lower = value.strip().lower()
    if lower in {"tbd", "dnp"}:
        return False
    if lower in _FORBIDDEN_OFFICIAL_SOURCE_VALUES:
        return True
    return any(phrase in lower for phrase in _FORBIDDEN_CUSTOMER_HEADER_PHRASES)


def _extract_explicit_official_source(line: BomLine) -> str | None:
    """Return canonical official source name only when explicitly present in notes."""
    text = (line.notes or "").strip()
    if not text:
        return None
    lower = text.lower()
    for needle, label in _OFFICIAL_SOURCE_PATTERNS:
        if needle in lower:
            return label
    # Exact match on notes when user entered a canonical source name directly.
    for _needle, label in _OFFICIAL_SOURCE_PATTERNS:
        if lower == label.lower():
            return label
    return None


def _resolve_official_row_from_snapshot(
    line: BomLine, snap_line: OfficialPriceLine | None
) -> tuple[str, float | None]:
    """Resolve official source + unit price from an official price snapshot line."""
    if snap_line is None:
        return "TBD", None
    source = snap_line.official_source or "TBD"
    if source.upper() == "DNP" or snap_line.pricing_status == "priced" and source == "DNP":
        return "DNP", None
    if snap_line.pricing_status == "missing_price" or source == "TBD":
        return "TBD", None
    if _is_forbidden_official_source(source):
        return "TBD", None
    if snap_line.official_unit_price is not None:
        return source, float(snap_line.official_unit_price)
    return source, None


def _resolve_official_row(line: BomLine) -> tuple[str, float | None]:
    """Resolve official source + unit price. No generic fallbacks; no price without source."""
    explicit = _extract_explicit_official_source(line)
    if explicit is None or _is_forbidden_official_source(explicit):
        return "TBD", None
    if line.customer_price is not None:
        return explicit, float(line.customer_price)
    return explicit, None


def _validate_official_source_values(values: list[str]) -> None:
    for value in values:
        if _is_forbidden_official_source(value):
            raise ValueError(
                f"Unsafe official source value '{value}' in customer export output"
            )


def _apply_data_row_number_formats(ws, row: int, *, is_dnp: bool) -> None:
    ws.cell(row=row, column=_COL_QTY).number_format = QTY_INT_FMT
    ws.cell(row=row, column=_COL_REQ_QTY).number_format = QTY_INT_FMT
    if not is_dnp:
        ws.cell(row=row, column=_COL_OFFICIAL_UNIT).number_format = USD_UNIT_FMT
    ws.cell(row=row, column=_COL_OFFICIAL_EXTENDED).number_format = USD_CURRENCY_FMT


def build_customer_bom_review_xlsx(
    db: Session,
    *,
    project: Project,
    version: BomVersion,
    lines: list[BomLine],
) -> tuple[bytes, str]:
    # Official/market reference export only — not final customer quote pricing.

    customer = db.get(Customer, project.customer_id)
    version_label = version.version_name or version.version_label
    file_name = customer_bom_review_filename(project.code, version_label)
    export_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    build_qty = version.build_quantity or project.build_quantity or 0

    wb, ws = _load_customer_template_sheet()
    template_extended_formula = _snapshot_template_extended_formula(ws)

    template_headers = [
        ws.cell(row=TEMPLATE_TABLE_HEADER_ROW, column=c).value
        for c in range(1, TEMPLATE_TABLE_COLS + 1)
    ]
    _validate_customer_export_headers([str(h) for h in template_headers if h])

    style_sources = _snapshot_row_styles(ws, TEMPLATE_DATA_START_ROW)

    _fill_customer_template_metadata(
        ws,
        customer_name=customer.name if customer else "—",
        project_name=project.name,
        project_code=project.code,
        version_label=version_label,
        revision=version.revision_code or "—",
        doc_number=version.source_doc_number or "—",
        board_name=version.board_name or "—",
        build_qty=build_qty,
        export_date=export_date,
    )

    last_data_row = TEMPLATE_DATA_START_ROW + len(lines) - 1 if lines else TEMPLATE_DATA_START_ROW
    data_end_row = max(last_data_row, TEMPLATE_DATA_DEFAULT_END_ROW)

    if last_data_row > TEMPLATE_DATA_DEFAULT_END_ROW:
        extra = last_data_row - TEMPLATE_DATA_DEFAULT_END_ROW
        ws.insert_rows(TEMPLATE_DATA_DEFAULT_END_ROW + 1, extra)
        data_end_row = last_data_row
        for offset in range(extra):
            new_row = TEMPLATE_DATA_DEFAULT_END_ROW + 1 + offset
            _apply_row_styles(ws, new_row, style_sources)
            if template_extended_formula:
                _apply_template_extended_formula(ws, new_row, template_extended_formula)

    _clear_bom_row_values(ws, TEMPLATE_DATA_START_ROW, data_end_row)

    official_snapshot = get_latest_exportable_snapshot(db, version.id)
    snapshot_by_line: dict[int, OfficialPriceLine] = {}
    if official_snapshot is not None:
        snapshot_by_line = snapshot_lines_by_bom_line(db, official_snapshot.id)

    written_sources: list[str] = []

    for row_offset, ln in enumerate(lines):
        row = TEMPLATE_DATA_START_ROW + row_offset
        is_dnp = _is_dnp_line(ln)

        if is_dnp:
            _apply_dnp_row_styles(ws, row, style_sources)
            official_source = "DNP"
            unit_price = None
            notes = _DNP_NOTE
        else:
            _apply_row_styles(ws, row, style_sources)
            snap_line = snapshot_by_line.get(ln.id)
            if snap_line is not None:
                official_source, unit_price = _resolve_official_row_from_snapshot(ln, snap_line)
            else:
                official_source, unit_price = _resolve_official_row(ln)
            notes = ln.notes

        written_sources.append(official_source)

        ws.cell(row=row, column=_COL_LINE, value=ln.line_no)
        ws.cell(row=row, column=_COL_MPN, value=ln.mpn)
        ws.cell(row=row, column=_COL_MANUFACTURER, value=ln.manufacturer)
        ws.cell(row=row, column=_COL_DESCRIPTION, value=ln.description)
        ws.cell(row=row, column=_COL_QTY, value=_num(ln.quantity))
        ws.cell(row=row, column=_COL_REQ_QTY, value=_num(ln.required_qty))
        ws.cell(row=row, column=_COL_REFDES, value=ln.reference_designators)
        ws.cell(row=row, column=_COL_FOOTPRINT, value=ln.footprint)
        ws.cell(row=row, column=_COL_OFFICIAL_SOURCE, value=official_source)
        ws.cell(row=row, column=_COL_OFFICIAL_UNIT, value=None if is_dnp else unit_price)
        ws.cell(row=row, column=_COL_NOTES, value=notes)

        ext_cell = ws.cell(row=row, column=_COL_OFFICIAL_EXTENDED)
        if is_dnp:
            ext_cell.value = 0
        elif not _cell_has_formula(ext_cell) and template_extended_formula:
            _apply_template_extended_formula(ws, row, template_extended_formula)

        _apply_data_row_number_formats(ws, row, is_dnp=is_dnp)

    _validate_official_source_values(written_sources)

    last_col = get_column_letter(TEMPLATE_TABLE_COLS)
    ws.auto_filter.ref = (
        f"A{TEMPLATE_TABLE_HEADER_ROW}:{last_col}{max(data_end_row, TEMPLATE_TABLE_HEADER_ROW)}"
    )
    if ws.freeze_panes is None:
        ws.freeze_panes = ws.cell(row=TEMPLATE_DATA_START_ROW, column=1)

    _validate_customer_export_headers(
        [
            str(ws.cell(row=TEMPLATE_TABLE_HEADER_ROW, column=c).value or "")
            for c in range(1, TEMPLATE_TABLE_COLS + 1)
        ]
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), file_name


def build_internal_bom_quality_xlsx(
    db: Session,
    *,
    project: Project,
    version: BomVersion,
    lines: list[BomLine],
) -> tuple[bytes, str]:
    from app.services.bom_line_override import quality_lines_for_version

    version_label = version.version_name or version.version_label
    file_name = internal_bom_quality_filename(project.code, version_label)
    quality_rows = quality_lines_for_version(db, version.id)

    wb, ws, tpl_ws = _create_internal_workbook("Internal BOM Quality")

    headers = [
        "Line",
        "Quality Status",
        "Effective MPN",
        "Uploaded MPN",
        "Manufacturer",
        "Uploaded Manufacturer",
        "Description",
        "Qty",
        "Required Qty",
        "RefDes",
        "Footprint",
        "DNP",
        "Needs Review",
        "Review Status",
        "Issues / Review Reason",
        "Review Notes",
        "Correction Note",
        "Has Correction",
        "Notes",
    ]
    data_rows = []
    for row in quality_rows:
        issues = row.get("review_reason") or ""
        data_rows.append(
            [
                row.get("line_number"),
                row.get("quality_status"),
                row.get("original_mpn"),
                row.get("uploaded_mpn"),
                row.get("manufacturer"),
                row.get("uploaded_manufacturer"),
                row.get("original_description"),
                row.get("qty_per_assembly"),
                row.get("required_qty"),
                row.get("reference_designators"),
                row.get("footprint"),
                "Yes" if row.get("is_dnp") else "No",
                "Yes" if row.get("needs_review") else "No",
                row.get("review_status"),
                issues,
                row.get("quality_review_note"),
                row.get("correction_note"),
                "Yes" if row.get("has_correction") else "No",
                row.get("notes"),
            ]
        )

    _render_internal_data_sheet(
        ws,
        report_type="bom_quality",
        meta_rows=[
            ("Project", project.name),
            ("Project Code", project.code),
            ("BOM Version", version_label),
            ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ],
        headers=headers,
        data_rows=data_rows,
        tpl_ws=tpl_ws,
    )
    _finalize_internal_workbook(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), file_name


def build_internal_pricing_snapshot_xlsx(
    db: Session,
    *,
    project: Project,
    snapshot: PricingSnapshot,
) -> tuple[bytes, str]:
    snap_name = snapshot.snapshot_name or snapshot.name
    file_name = internal_pricing_snapshot_filename(project.code, snap_name)

    p_lines = list(
        db.scalars(
            select(PricingLine)
            .where(PricingLine.pricing_snapshot_id == snapshot.id)
            .order_by(PricingLine.id)
        )
    )
    bom_by_id: dict[int, BomLine] = {}
    bom_ids = [pl.bom_line_id for pl in p_lines if pl.bom_line_id is not None]
    if bom_ids:
        for bl in db.scalars(select(BomLine).where(BomLine.id.in_(bom_ids))):
            bom_by_id[bl.id] = bl

    wb, ws, tpl_ws = _create_internal_workbook("Internal Pricing")

    headers = [
        "Line",
        "MPN",
        "Manufacturer",
        "Description",
        "Required Qty",
        "Unit Cost",
        "Extended Cost",
        "Currency",
        "Pricing Status",
        "Match Confidence",
        "Selected Source",
        "Notes",
    ]
    data_rows = []
    for idx, pl in enumerate(p_lines, start=1):
        bl = bom_by_id.get(pl.bom_line_id) if pl.bom_line_id else None
        data_rows.append(
            [
                bl.line_no if bl else idx,
                pl.mpn or (bl.mpn if bl else None),
                bl.manufacturer if bl else None,
                bl.description if bl else None,
                _num(pl.required_qty),
                _num(pl.unit_cost),
                _num(pl.extended_cost),
                pl.currency or snapshot.currency,
                pl.pricing_status,
                pl.match_confidence,
                pl.selected_source,
                pl.notes,
            ]
        )

    _render_internal_data_sheet(
        ws,
        report_type="pricing_snapshot",
        meta_rows=[
            ("Project", project.name),
            ("Project Code", project.code),
            ("Pricing Snapshot", snap_name),
            ("Currency", snapshot.currency),
        ],
        headers=headers,
        data_rows=data_rows,
        tpl_ws=tpl_ws,
    )
    _finalize_internal_workbook(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), file_name


def internal_pricing_workbench_filename(project_code: str, version_name: str) -> str:
    return f"Internal_Pricing_Workbench_{_safe_part(project_code)}_{_safe_part(version_name)}.xlsx"


def internal_pricing_comparison_filename(project_code: str, version_name: str) -> str:
    return f"Internal_Pricing_Comparison_{_safe_part(project_code)}_{_safe_part(version_name)}.xlsx"


def supplier_purchase_report_filename(
    project_code: str, version_name: str, supplier: str
) -> str:
    return (
        f"Supplier_Purchase_{_safe_part(supplier)}_{_safe_part(project_code)}"
        f"_{_safe_part(version_name)}.xlsx"
    )


def _pricing_mode_label(include_east: bool) -> str:
    return "משולב עם מחירי מזרח" if include_east else "רשמי בלבד"


def _offers_summary(offers: list[dict] | None) -> str:
    if not offers:
        return ""
    parts: list[str] = []
    for o in offers:
        label = o.get("supplier_display") or o.get("supplier") or "?"
        price = o.get("unit_price")
        parts.append(f"{label}: {price if price is not None else '—'}")
    return " | ".join(parts)


def _offer_label(offer: dict | None) -> str | None:
    if not offer:
        return None
    return offer.get("supplier_display") or offer.get("supplier")


def _best_offer_by_internal(offers: list[dict] | None, internal: bool) -> dict | None:
    if not offers:
        return None
    pool = [o for o in offers if bool(o.get("internal_only")) == internal]
    if not pool:
        return None
    priced = [o for o in pool if o.get("unit_price") is not None]
    if not priced:
        return pool[0]
    return min(priced, key=lambda o: float(o["unit_price"]))


def _workbench_data(
    db: Session,
    *,
    project: Project,
    version: BomVersion,
    include_east: bool,
) -> dict:
    from app.services.suppliers.workbench import get_workbench_results

    return get_workbench_results(
        db,
        project_id=project.id,
        bom_version_id=version.id,
        include_east_override=include_east,
    )


def build_internal_pricing_workbench_xlsx(
    db: Session,
    *,
    project: Project,
    version: BomVersion,
    include_east: bool | None = None,
) -> tuple[bytes, str]:
    version_label = version.version_name or version.version_label
    east_on = (
        bool(include_east)
        if include_east is not None
        else bool(version.include_east_pricing)
    )
    file_name = internal_pricing_workbench_filename(project.code, version_label)
    data = _workbench_data(db, project=project, version=version, include_east=east_on)
    cmp_data = data.get("pricing_comparison") or {}

    wb, ws, tpl_ws = _create_internal_workbook("Pricing Lines")

    off = cmp_data.get("official_only") or {}
    east = cmp_data.get("with_east") or {}
    savings = cmp_data.get("savings") or {}

    headers = [
        "Line",
        "MPN",
        "Manufacturer",
        "Description",
        "Required Qty",
        "Selected Source",
        "Source Type",
        "Supplier PN",
        "Unit Price",
        "Extended Price",
        "Stock",
        "Lead Time",
        "Status",
        "Solution Status",
        "Internal Only",
        "Official Best Ext",
        "East Best Ext",
        "Savings",
        "Available Offers",
        "Notes",
    ]
    data_rows = []
    for ln in data.get("lines") or []:
        lp = ln.get("line_pricing") or {}
        offers = ln.get("offers") or []
        data_rows.append(
            [
                ln.get("line_no"),
                ln.get("mpn"),
                ln.get("manufacturer"),
                ln.get("description"),
                ln.get("required_qty"),
                ln.get("source"),
                ln.get("selected_source_type"),
                ln.get("supplier_part_number"),
                ln.get("unit_price"),
                ln.get("extended_price"),
                ln.get("stock"),
                ln.get("lead_time"),
                ln.get("status"),
                ln.get("solution_status"),
                "Yes" if ln.get("source_is_internal") else "No",
                lp.get("official_best_extended"),
                lp.get("east_best_extended"),
                lp.get("difference"),
                _offers_summary(offers),
                ln.get("notes"),
            ]
        )

    _render_internal_data_sheet(
        ws,
        report_type="pricing_snapshot",
        meta_rows=[
            ("Project", project.name),
            ("Project Code", project.code),
            ("BOM Version", version_label),
            ("Pricing Mode", _pricing_mode_label(east_on)),
        ],
        kpi_rows=[
            ("Official Only Total", off.get("total")),
            ("With East Total", east.get("total")),
            ("Savings Amount", savings.get("amount")),
            ("Savings Percent", savings.get("percent")),
        ],
        headers=headers,
        data_rows=data_rows,
        tpl_ws=tpl_ws,
    )
    _finalize_internal_workbook(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), file_name


def build_internal_pricing_comparison_xlsx(
    db: Session,
    *,
    project: Project,
    version: BomVersion,
    include_east: bool | None = None,
) -> tuple[bytes, str]:
    version_label = version.version_name or version.version_label
    east_on = (
        bool(include_east)
        if include_east is not None
        else bool(version.include_east_pricing)
    )
    file_name = internal_pricing_comparison_filename(project.code, version_label)
    data = _workbench_data(db, project=project, version=version, include_east=east_on)
    cmp_data = data.get("pricing_comparison") or {}

    wb, ws_sum, tpl_ws = _create_internal_workbook("Summary")
    off = cmp_data.get("official_only") or {}
    east = cmp_data.get("with_east") or {}
    savings = cmp_data.get("savings") or {}
    summary = data.get("summary") or {}

    _render_internal_summary_sheet(
        ws_sum,
        report_type="summary",
        meta_rows=[
            ("Project", project.name),
            ("BOM Version", version_label),
            ("Pricing Mode", _pricing_mode_label(east_on)),
            ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Lines", summary.get("total_lines")),
            ("Has Solution", summary.get("has_solution")),
            ("Needs Approval", summary.get("needs_approval")),
            ("No Solution", summary.get("no_solution")),
            ("DNP", summary.get("dnp")),
            ("Official Only Total", off.get("total")),
            ("With East Total", east.get("total")),
            ("Savings Amount", savings.get("amount")),
            ("Savings Percent", savings.get("percent")),
        ],
        tpl_ws=tpl_ws,
    )

    ws_lines = _add_internal_sheet(wb, "Line Comparison", tpl_ws)
    headers = [
        "Line",
        "MPN",
        "Required Qty",
        "Selected Source",
        "Selected Extended",
        "Official Best Source",
        "Official Best Ext",
        "East Best Source",
        "East Best Ext",
        "Savings",
        "Solution Status",
    ]
    line_rows = []
    for ln in data.get("lines") or []:
        lp = ln.get("line_pricing") or {}
        offers = ln.get("offers") or []
        off_offer = _best_offer_by_internal(offers, internal=False)
        east_offer = _best_offer_by_internal(offers, internal=True)
        line_rows.append(
            [
                ln.get("line_no"),
                ln.get("mpn"),
                ln.get("required_qty"),
                ln.get("source"),
                ln.get("extended_price"),
                _offer_label(off_offer),
                lp.get("official_best_extended"),
                _offer_label(east_offer),
                lp.get("east_best_extended"),
                lp.get("difference"),
                ln.get("solution_status"),
            ]
        )
    _render_internal_data_sheet(
        ws_lines,
        report_type="pricing_comparison",
        meta_rows=[
            ("Project", project.name),
            ("BOM Version", version_label),
            ("Pricing Mode", _pricing_mode_label(east_on)),
            ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ],
        headers=headers,
        data_rows=line_rows,
        tpl_ws=tpl_ws,
    )

    _finalize_internal_workbook(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), file_name


_PURCHASE_SUPPLIER_FILTERS = frozenset(
    {"all", "digikey", "mouser", "ti", "link", "manual", "tbd"}
)


def _normalize_purchase_supplier_filter(value: str | None) -> str:
    if not value:
        return "all"
    key = value.strip().lower().replace("-", "").replace(" ", "")
    mapping = {
        "all": "all",
        "כלהספקים": "all",
        "digikey": "digikey",
        "mouser": "mouser",
        "ti": "ti",
        "link": "link",
        "east": "link",
        "manual": "manual",
        "tbd": "tbd",
        "nosolution": "tbd",
    }
    if key in mapping:
        return mapping[key]
    if key in _PURCHASE_SUPPLIER_FILTERS:
        return key
    return "all"


def _purchase_line_supplier_key(line: dict) -> str:
    st = line.get("selected_source_type")
    if st == "dnp":
        return "dnp"
    if st == "manual":
        return "manual"
    if st == "tbd" or line.get("solution_status") == "No Solution":
        return "tbd"
    if st == "east_quote" or line.get("source_is_internal"):
        return "link"
    supplier = (line.get("selected_supplier") or "").lower()
    if supplier in ("digikey", "mouser", "ti", "link"):
        return supplier
    src = (line.get("source") or "").lower()
    if "digi" in src:
        return "digikey"
    if "mouser" in src:
        return "mouser"
    if src == "ti" or "texas instruments" in src:
        return "ti"
    if "link" in src:
        return "link"
    if line.get("source") == "TBD":
        return "tbd"
    return "manual"


def _line_in_purchase_report(line: dict, supplier_filter: str, include_east: bool) -> bool:
    key = _purchase_line_supplier_key(line)
    if key == "dnp":
        return False
    if not include_east and key == "link":
        return False
    if supplier_filter == "all":
        if key == "tbd":
            return True
        return key != "dnp"
    if supplier_filter == "tbd":
        return key == "tbd"
    return key == supplier_filter


def _purchase_row_from_line(line: dict) -> list:
    offers = line.get("offers") or []
    match_reason = ""
    for o in offers:
        if o.get("is_currently_selected"):
            match_reason = o.get("match_reason") or ""
            break
    return [
        line.get("source"),
        "Yes" if line.get("source_is_internal") else "No",
        line.get("mpn"),
        line.get("manufacturer"),
        line.get("description"),
        line.get("reference_designators"),
        line.get("required_qty"),
        line.get("supplier_part_number"),
        line.get("unit_price"),
        line.get("extended_price"),
        line.get("currency") or "USD",
        line.get("stock"),
        line.get("lead_time"),
        line.get("status"),
        line.get("solution_status"),
        line.get("notes"),
        match_reason,
    ]


_PURCHASE_LINE_HEADERS = [
    "Supplier / Source",
    "Internal Only",
    "MPN",
    "Manufacturer",
    "Description",
    "Designators",
    "Required Qty",
    "Supplier PN",
    "Unit Price",
    "Extended Price",
    "Currency",
    "Stock / Available Qty",
    "Lead Time",
    "Status",
    "Solution Status",
    "Notes",
    "Match Reason",
]


def build_supplier_purchase_report_xlsx(
    db: Session,
    *,
    project: Project,
    version: BomVersion,
    supplier_filter: str = "all",
    include_east: bool = False,
) -> tuple[bytes, str]:
    version_label = version.version_name or version.version_label
    filt = _normalize_purchase_supplier_filter(supplier_filter)
    file_name = supplier_purchase_report_filename(project.code, version_label, filt)
    data = _workbench_data(db, project=project, version=version, include_east=include_east)

    purchase_rows: list[dict] = []
    supplier_totals: dict[str, float] = {}
    dnp_count = 0
    tbd_count = 0
    needs_approval = 0
    no_stock = 0

    for ln in data.get("lines") or []:
        key = _purchase_line_supplier_key(ln)
        if key == "dnp":
            dnp_count += 1
            continue
        if ln.get("solution_status") == "Needs Approval":
            needs_approval += 1
        if ln.get("status") == "No Stock":
            no_stock += 1
        if key == "tbd":
            tbd_count += 1
        if not _line_in_purchase_report(ln, filt, include_east):
            continue
        purchase_rows.append(ln)
        ext = ln.get("extended_price")
        if ext is not None and ln.get("solution_status") != "DNP":
            sk = key if key != "tbd" else "TBD"
            supplier_totals[sk] = supplier_totals.get(sk, 0.0) + float(ext)

    grand_total = sum(supplier_totals.values())

    wb, ws_sum, tpl_ws = _create_internal_workbook("Summary")

    _render_internal_summary_sheet(
        ws_sum,
        report_type="summary",
        meta_rows=[
            ("Project", project.name),
            ("Project Code", project.code),
            ("BOM Version", version_label),
            ("Pricing Mode", _pricing_mode_label(include_east)),
            ("Supplier Filter", filt),
            ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Lines", len(data.get("lines") or [])),
            ("Purchase Lines", len(purchase_rows)),
            ("DNP Excluded", dnp_count),
            ("TBD / No Solution", tbd_count),
            ("Needs Approval", needs_approval),
            ("No Stock", no_stock),
            ("Grand Total", grand_total),
        ],
        section_title="סה״כ לפי ספק / Supplier Totals",
        section_rows=[(f"Total — {supplier}", total) for supplier, total in sorted(supplier_totals.items())],
        tpl_ws=tpl_ws,
    )

    ws_lines = _add_internal_sheet(wb, "Purchase Lines", tpl_ws)
    _render_internal_data_sheet(
        ws_lines,
        report_type="supplier_purchase",
        meta_rows=[
            ("Project", project.name),
            ("BOM Version", version_label),
            ("Supplier Filter", filt),
            ("Pricing Mode", _pricing_mode_label(include_east)),
        ],
        headers=_PURCHASE_LINE_HEADERS,
        data_rows=[_purchase_row_from_line(ln) for ln in purchase_rows],
        tpl_ws=tpl_ws,
    )

    sheet_map = {
        "digikey": "Digi-Key",
        "mouser": "Mouser",
        "ti": "TI",
        "link": "Link",
        "manual": "Manual",
        "tbd": "TBD",
    }
    if filt == "all":
        by_supplier: dict[str, list[dict]] = {}
        for ln in purchase_rows:
            sk = _purchase_line_supplier_key(ln)
            by_supplier.setdefault(sk, []).append(ln)
        for sk, title in sheet_map.items():
            supplier_rows = by_supplier.get(sk)
            if not supplier_rows:
                continue
            ws = _add_internal_sheet(wb, title, tpl_ws)
            _render_internal_data_sheet(
                ws,
                report_type="supplier_purchase",
                meta_rows=[
                    ("Project", project.name),
                    ("BOM Version", version_label),
                    ("Supplier", title),
                    ("Pricing Mode", _pricing_mode_label(include_east)),
                ],
                headers=_PURCHASE_LINE_HEADERS,
                data_rows=[_purchase_row_from_line(ln) for ln in supplier_rows],
                tpl_ws=tpl_ws,
            )

    _finalize_internal_workbook(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), file_name


def build_supplier_pricing_workbench_xlsx(
    db: Session,
    *,
    project: Project,
    version: BomVersion,
) -> tuple[bytes, str]:
    from app.services.suppliers.workbench import get_workbench_results

    version_label = version.version_name or version.version_label
    file_name = supplier_workbench_filename(project.code, version_label)
    data = get_workbench_results(
        db, project_id=project.id, bom_version_id=version.id
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Supplier Pricing"
    ws.sheet_view.rightToLeft = True

    headers = [
        "MPN",
        "Manufacturer",
        "Description",
        "Required Qty",
        "Source",
        "Supplier PN",
        "Unit Price",
        "Extended Price",
        "Stock",
        "Status",
        "Solution Status",
        "Notes",
    ]
    forbidden = [
        "china",
        "internal",
        "margin",
        "savings",
        "match confidence",
        "buyer",
    ]
    for h in headers:
        if any(f in h.lower() for f in forbidden):
            raise ValueError(f"Unsafe workbench export header '{h}'")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for ln in data["lines"]:
        ws.append(
            [
                ln.get("mpn"),
                ln.get("manufacturer"),
                ln.get("description"),
                ln.get("required_qty"),
                ln.get("source"),
                ln.get("supplier_part_number"),
                ln.get("unit_price"),
                ln.get("extended_price"),
                ln.get("stock"),
                ln.get("status"),
                ln.get("solution_status"),
                ln.get("notes"),
            ]
        )

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), file_name
